from bs4 import BeautifulSoup as soup
import requests 
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from openpyxl import Workbook,load_workbook
from datetime import datetime,timedelta
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import json
import re
from helpers.db_helpers import *
from time import sleep
from random import randint
from config import yt_api_key
#from translateAPI import my_translate,detect_language
# from Dicts_Main import langCode

####---------INPUTS HERE-------------------------------------
####---------INPUTS HERE-------------------------------------
output_path ='output/'
replace_sheet = 'Replace List'

youtube_input_path="helpers/YoutubeMagicRSS.xlsx"
yt_rss_base = "https://www.youtube.com/feeds/videos.xml?channel_id="
yt_api_base_videos = "https://youtube.googleapis.com/youtube/v3/videos?part=contentDetails&part=snippet&id="
yt_api_base_channels = "https://www.googleapis.com/youtube/v3/channels?"

####----------END INPUTS----------------------------------------

####----GETTING THE CHANNELS FROM DB-----------------------------
db_yt_sources_dict = best_sources_channels()
db_yt_sources_ids = list(best_sources_channels().keys())
explore_sources_dict = explore_channels()
all_sources_dict = {**db_yt_sources_dict, **explore_sources_dict}
db_explore_channels_ids = list(explore_sources_dict)
all_ids = db_yt_sources_ids + db_explore_channels_ids

####----------Boundaries for filters--------------------
tooLong = timedelta(minutes=9) #10 minutes - max length of a video
tooShort = timedelta(seconds=60)#1,5 minutes - min length of a video
tooOld = 60 #days since a video was published
ytCategory = 10#10 = Music, but sometimes it is only music-related
###----------------------------------


date_stamp=datetime.now()
excel_date = date_stamp.strftime('%d.%m.%Y')
my_date_stamp = date_stamp.strftime('%y%m%d_%H-%M')
excel_sheet_date = date_stamp.strftime('%d %M')
some_path = f"/Users/viktorkertanov/Downloads/YoutubeRSS_{my_date_stamp}.xlsx"

###---------------START-Youtube----DATE----TIME---CONVERSIONS----FUNCTIONS
#https://en.wikipedia.org/wiki/ISO_8601#Durations for duration

def yt_duration(yt_dur):
    duration_tuple = re.findall("P(\d+D)?T?(\d+H)?(\d+M)?(\d+S)?", yt_dur)[0]
    dur_lst = []
    for el in duration_tuple:
        el = re.split("\D", el)[0]
        if not el:
            el = 0
        dur_lst.append(int(el))
    duration = timedelta(days = dur_lst[0], hours = dur_lst[1], minutes = dur_lst[2], seconds = dur_lst[3])
    return duration

def yt_time(yt_time):
   try:
      time = datetime.strptime(yt_time, '%Y-%m-%dT%H:%M:%S%z')
      time = time.replace(tzinfo=None)#deleting UTC offset so we can calculate how old the video is
   except:
      time = yt_time
   return time
###---------------END-Youtube----DATE----TIME---CONVERSIONS----FUNCTIONS
###-----function that divides a list into evenly n-sized chunks; the remainder goes to the last chunk
def even_chunks(lst, n):  
   for i in range(0, len(lst), n):
      yield lst[i:i + n]

#-------Functions that work with strings and modify them using several methods---------------------


def white_title(string):#f(x) for deleting traling and excessive whitespace and if the title is all low or all UPPER changing it to proper title
    string= re.sub(' +',' ',string).strip()
    if string.split('(')[0].islower() or string.split('(')[0].isupper():
        string = string.title()
    return string


def big_split(string,channel,split_chart):#splitting the string based on the rules for certain YT channels
    for el in split_chart:
        if channel == el[0]:
            try:
                string = string.split(el[1])[el[2]]
            except:
                string = string
            break
##    print(string)
    return string


def regex_repl(string,regex_chart):#replacing using regex chart
    for rgx in regex_chart:
        string = re.sub(rgx[0],rgx[1],string)
##    print(string)
    return string
def big_replace(string, replace_chart):#replacing one string with another using the replace chart
    mod_repl_chart = []
    for repl_pair in replace_chart:
        if len(mod_repl_chart)!=0:
            for mod_el in mod_repl_chart:
                repl_pair[0] = repl_pair[0].replace(mod_el[0],mod_el[1])
        mod_repl_chart.append(repl_pair)
        string = string.replace(repl_pair[0],repl_pair[1])
    return string


def del_emoji(string):#f(x) for deleting emojis
    EMOJI_PATTERN = re.compile(
    "(["
    "\U0001F1E0-\U0001F1FF"  # flags (iOS)
    "\U0001F300-\U0001F5FF"  # symbols & pictographs
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F680-\U0001F6FF"  # transport & map symbols
    "\U0001F700-\U0001F77F"  # alchemical symbols
    "\U0001F780-\U0001F7FF"  # Geometric Shapes Extended
    "\U0001F800-\U0001F8FF"  # Supplemental Arrows-C
    "\U0001F900-\U0001F9FF"  # Supplemental Symbols and Pictographs
    "\U0001FA00-\U0001FA6F"  # Chess Symbols
    "\U0001FA70-\U0001FAFF"  # Symbols and Pictographs Extended-A
    "\U00002702-\U000027B0"  # Dingbats
    "])"
    )
    string = re.sub(EMOJI_PATTERN,' ', string)
    return string


def feat_extractor(string):
    feat = " Feat. "
    try:
        track = white_title(string.split(' - ')[1])
##        print(track)
        artist = string.split(' - ')[0]
##        print(artist)
        if feat in track:
            feat_meta = f" Feat. {track.split(feat)[1]}"
##            print(feat_meta)
            artist +=feat_meta
##            print(artist)
            track = track.replace(feat_meta,'')
            track = white_title(track)
##            print(track)
        overall = f"{artist} - {track}"
    except:
        track = string
        artist = string
        overall = string
    return white_title(track), artist, overall


def global_replace(string,channel=None):#using the sequence of all the replace functions that we have
    string = del_emoji(string)#1) deleting all the un-neccessary emojis from the string
    if channel!=None:
        string = big_split(string,channel,split_chart)
    string = regex_repl(string,regex_chart)
    string = big_replace(string,replace_chart)
    string = white_title(string)
    return string


def fresh_status(string):
    if "now" not in string:
        days = int(re.sub('[a-zA-Z]','',string))
    if "min" in string or "h" in string or "now" in string:
        status = f"HOT!: {string}"
    elif days <5:
        status = f"NEW!: {string}"
    elif days < 15:
        status = f"OK: {string}"
    else:
        status = f"OLD: {string}"
    return status


## ----------------------------------------------------------


def excel_extractor(path, sheet_name, max_column=None, min_column=None, headers=False):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row
    if max_column == None:
        max_column = ws.max_column
    else:
        max_column = max_column
    result_chart = []
    if headers == False:
        min_row = 2
    else:
        min_row = 1#we extract the data tha has no headers
    if min_column == None:
        min_column = 1
    else:
        min_column = min_column
    for row in ws.iter_rows(min_row=min_row, min_col=min_column, max_col=max_column, max_row=max_row, values_only=True):
        total_val = 0
        for cell in row:
            if cell!=None:
                total_val+=len(str(cell))
                try:
                    url = cell.hyperlink.target
                except:
                    url = ''
        if total_val>0:
            result_chart.append(list(row))
    return result_chart

##-------END----FUNCTIONS----FOR-READ-WRITE-EXCELS-------------------------------

###------EXTRACTING--DATA---FOR----TITLE-----CONVERSIONS
input_replace_chart = excel_extractor(youtube_input_path,replace_sheet)#raw input replace-split chart from Excel

split_chart,regex_chart,replace_chart = [],[],[]

for row in input_replace_chart:#changing one column cells from None to '' in order to create the proper replace
    if not row[2]:
        row[2] = ''

for el in input_replace_chart:#breakdown of Excel input replace chart into 3) main parts: split by YT channel rules; regex replace; standard replace
    if el[4] == "CHANNEL":
        split_row = [el[0],el[1],el[3]]
        split_chart.append(split_row)#creating the chart for the big_split f(x)
    elif el[4] == "REGEX":
        regex_row = [el[1],el[2]]
        regex_chart.append(regex_row)#creating the chart for the regex_repl f(x)
    else:
        replace_row = [el[1],el[2]]
        replace_chart.append(replace_row)#creating the chart for the big_replace f(x)


final_xml_chart = {}
translate_xml_chart = {}
counter=0
min_cnl_index = 0 # we can set where in the list we'll be diving
max_cnl_index = 3000 # we can set the max row in all channels

def cover_remix_live(song_title_string):
    cover_remix_live = ['live', 'remix','cover']
    result = 0
    for el in cover_remix_live:
        songTitle = song_title_string.upper()
        elUpper = el.upper()
        if elUpper in songTitle:
            result = el
            return result
        else:
            continue
    if result == 0:
        return None
stopWordsExcel = excel_extractor(youtube_input_path, 'stopWords', max_column=1)
stopWordsList = [cell.upper() for row in stopWordsExcel for cell in row]

def stopWords(song_title_string):
    songTitleUpper = song_title_string.upper()
    stopWordStatus = None
    for el in stopWordsList:
        if el in songTitleUpper:
            stopWordStatus = 'STOP WORD'
            return stopWordStatus
        else:
            continue
    if stopWordStatus == None:
        return None

if max_cnl_index>len(all_ids):
   max_cnl_index=len(all_ids)

cnl_index = 1
for channel in all_ids[min_cnl_index:max_cnl_index]:
   sleep(randint(1, 3))
   print(f"""{cnl_index} out of {max_cnl_index-min_cnl_index}. We are doing this channel: {channel} ::: {(', ').join([str(i) for i in all_sources_dict[channel]])}""")
   cnl_index += 1
   with requests.Session() as session:
      # try:
     rss_url = yt_rss_base+channel
     try:
        myClient = urlopen(rss_url)
     except HTTPError as e:
        print(f"""Error code: {e.code} {e.reason}. URL:{rss_url} (see above)""")
        continue
     except URLError as e:
        print('Reason: ', e.reason)
        continue
     page_html = myClient.read()
     myClient.close()
     page_soup = soup(page_html,"html.parser")
     videos = page_soup.findAll('entry')
     for video in videos:
        videoID = video.find("yt:videoid").text
        channelID = video.find("yt:channelid").text
        videoTitle = video.find("media:title").text
        videoLink = video.find("link").get("href")
        channelName = video.find("name").text
        channelLink = video.find("uri").text
        published = yt_time(video.find("published").text)
        description = video.find("media:description").text
        viewsCount = int(video.find("media:statistics").get("views"))
        avgRating = round(float(video.find("media:starrating").get("average")), 1)
        coverRemixLive = cover_remix_live(videoTitle)
        stopWord = stopWords(videoTitle)

        timeDelta = (date_stamp - published).days
        new_videoTitle = global_replace(videoTitle, channelName)#working with original metadata string
        track, artist, new_videoTitle = feat_extractor(new_videoTitle)#obtaining the format for batch search

        if timeDelta<=tooOld:
           counter+=1#counting the crows, i.e. rows
           # print(f"We are adding the video # {counter}")
           final_xml_chart[videoID] = [
              videoID,
              channelID,
              channelLink,
              channelName,
              videoTitle,
              new_videoTitle,
              track, '|', artist,
              '','',#for personal checks of the file
              timeDelta,
              published,
              f"{artist} - {track}",
              videoLink,
              viewsCount,
              avgRating,
              coverRemixLive,
              stopWord,
              description
           ]

           translate_xml_chart[videoID] = [videoTitle, description]#chart for creating translations
        else:
##               print(f"We have not included this track:{videoID},it was published {timeDelta} days ago")
           continue
      # except:
      #    print(f"Some problem with the channel: {channel} ::: {(', ').join(db_yt_sources_dict[channel])}")
      #    continue#not the best catch in the world. temporary

#now we are obtaining a list of videoIDs to connect to youtube API for getting the duration and category

videoID_list = list(final_xml_chart.keys())
chunk_size=50
api_chunks = list(even_chunks(videoID_list,chunk_size))
yt_api_dict = {}
for chunk in api_chunks:
   chunk = (",").join(chunk)
   api_query = yt_api_base_videos+chunk+"&key="+yt_api_key
   json_response = requests.get(api_query)
   response = json.loads(json_response.text)
   response_videos = response["items"]
   for item in response_videos:
      item_id = item["id"]
      duration = item["contentDetails"]["duration"]
      duration = yt_duration(duration)
      # duration = duration
      category_ID = int(item["snippet"]["categoryId"])
      try:
         if category_ID == 10 and duration >= tooShort and duration <= tooLong:
            item_row = [duration,category_ID]
            yt_api_dict[item_id] = item_row
         else:
            print(f"We have excluded this, because: cat is {category_ID}, duration is {duration}")
            del final_xml_chart[item_id]
      except:
         item_row = [duration,category_ID]
         yt_api_dict[item_id] = item_row
         print(f"we are in an unknown exception: duration is {duration}, yt cat is {category_ID}")

#adding duration + category info to the final chart
for elem in yt_api_dict:
   if elem in final_xml_chart:#we add additional info only if the videoID(=elem) is already in the final xml chart
      additional_info = yt_api_dict[elem]
      final_xml_chart[elem] = final_xml_chart[elem]+additional_info
   else:
      print(f"we are somehow here:{elem}")
      continue
##    print(final_xml_chart[elem])

#### --------------------FILTER SECTION: it is not probably optimal to put it here, but it is more efficient as you can always change and get the FULL data
##final_xml_chart_filtered = []
##for element in final_xmal_chart:
##   

####--------------------------------------------------------------
#compiling the final chart, structured as the dict (key=videoID): info from RSS xml + additional info from youtube API
excel_headers = [
   'videoID',
   'channelID',
   'channelURL',
   'channelName',
   'Original Metadata',
   'NEW Metadata',
   'Track','|','Artist',
   'Check','Upload',
   'DaysOld',
   'Publish Date',
   'Artist - Track',
   'Video URL',
   'Views',
   'Rating',
   'Cover-Rmx-Live',
   'StopWord',
   'Original Description',
   'Length',
   'YT Category'
]
#Костыль: стоп-слова
stop_list = [
    'karaoke',
    'tutorial',
    'making of',
    'lesson',
    'interview',
    'reaction',
    'react',
    'type beat',
    '(FREE)',
    'backstage'
]

duplicates_delete = processed_videos_list()

output_chart = []
duplicate_counter = 0
for key in final_xml_chart:
    yt_video_id = final_xml_chart[key][0]
    original_meta = final_xml_chart[key][4]
    if yt_video_id not in duplicates_delete:
        output_chart.append(final_xml_chart[key])
    else:
        duplicate_counter += 1
        # print(f"We have already processed this track id: {yt_video_id}")


#writing the chart to the path&sheet_name
def excel_writer(headers,chart,sheet_name='YoutubeMagic'):
   wb = Workbook()
   ws = wb.active
   ws.title = f'{sheet_name}_{excel_sheet_date}'
   ws.append(headers)
   for row in chart:
     ws.append(row)
   if sheet_name == 'YoutubeMagic':   
     #Formatting the final chart
     ##------sheet1--------
     ws = wb.active
     ws.sheet_properties.tabColor = "90EE90"
   width_dict = {
   2:["H"],

   9:["J","K","L","P","Q","R","U","V"],

   20:["M"],

   25:["A","B","C","D"],

   42:["G","I","N","O","S","T"],

   70:["E","F"]}

   for width in width_dict:
      for column in width_dict[width]:
         ws.column_dimensions[column].width = width
   
   thin = Side(border_style="thin", color="000000")
   thick = Side(border_style="thick",color="000000")

   ws_workrange=ws['A2:V'+str(len(chart)+1)]
   ws_headers=ws['A1:V1']

   ws.auto_filter.ref ='A1:V'+str(len(chart)+1)
   for element in ws_workrange:
      for cell in element:
          cell.alignment = Alignment(horizontal='left',vertical='center')
   for element in ws_headers:
      for cell in element:
          cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
          cell.fill = PatternFill('solid',fgColor="DDDDDD")
          cell.font = Font(bold=True)
          cell.alignment = Alignment(horizontal='center',vertical='center')                             
   # ws.column_dimensions.group('A','D',hidden=True)
   # ws.column_dimensions.group('N','O',hidden=True)
   # ws.column_dimensions.group('S','T',hidden=True)
   wb.save(f'{output_path}{sheet_name}_{my_date_stamp}.xlsx')
   return None

excel_writer(excel_headers,output_chart)
