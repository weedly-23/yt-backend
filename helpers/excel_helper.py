from openpyxl import Workbook, load_workbook

downloadsFolder = "/output/"



def excel_reader(path, sheet_name, max_column=None, min_column=None, max_row=None, include_headers=True):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    final_chart = []
    if max_column == None:
        max_column = ws.max_column
    if min_column == None:
        min_column = ws.min_column
    else:
        min_column = min_column
    if include_headers == False:
        min_row = 2
    else:
        min_row = 1
    if max_row == None:
        max_row = ws.max_row
    else:
        max_row = max_row
    for row in ws.iter_rows(min_row = min_row, min_col=min_column, max_col=max_column, max_row=max_row, values_only=True):
        total_val = 0
        for cell in row:
            if cell!=None:
                total_val+=len(str(cell))
                try:
                    url = cell.hyperlink.target
                except:
                    url = ""
        if total_val>0:
            final_chart.append(list(row))
    return final_chart

def list_2_dict(input_list, dict_key_index=0, dict_key_name=None, dict_key_name_duplicate_choicemaker=None):
    final_dict = {}
    key_names_list = input_list[0]#We take first row, i.e. headers as the NAME of the keys in final DICT
    if not dict_key_name:
        external_key = key_names_list[dict_key_index]
    else:
        if dict_key_name in key_names_list:
            external_key = dict_key_name
        else:
            print(f"We don't have that key - [{dict_key_name}] - in the headers.\n"
                  f"Try one of these keys: {(', '.join(key_names_list))}")
            return None
        if dict_key_name_duplicate_choicemaker in key_names_list:
            choicemaker_key = dict_key_name_duplicate_choicemaker
        else:
            print(f"We don't have the key you suggested for making choices between duplicates: {dict_key_name_duplicate_choicemaker}\n"
                  f"Try one of these keys: {(', '.join(key_names_list))}")
            return None
    data_chart = input_list[1:]
    for row in data_chart:
        internal_row_dict = {key_names_list[i]: row[i] for i in range(len(key_names_list))}
        current_key = internal_row_dict[external_key]
        if current_key not in final_dict:
            final_dict[current_key] = internal_row_dict
        elif final_dict[current_key][choicemaker_key] < internal_row_dict[choicemaker_key]:
            final_dict[current_key] = internal_row_dict
            print(f"It seems like there's a duplicate in the original chart: {current_key}")
        else:
            continue
    return final_dict


def excel_writer(output_list, filename, folder_path=downloadsFolder, sheet_name=None):
    wb = Workbook()
    if sheet_name != None:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    for row in output_list:
        ws.append(row)
    wb.save(f"{folder_path}{filename}.xlsx")
    return None