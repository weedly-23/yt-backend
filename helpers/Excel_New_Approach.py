from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, GradientFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import utils
from helpers.excel_helper import excel_reader, excel_writer
from datetime import datetime
from helpers.aux_lists import Colors
import random

colors = Colors().pallete


zeroFill = colors["Cool Gray 2"]
oneFill = colors["Chart Green 2"]
twoFill = colors["Chart Orange 2"]
threeFill = colors["Chart Purple 2"]
riskFill = "E7B1B5"
medRiskFill = "F6D390"

date_stamp = datetime.now()
today = date_stamp.date()  # date in date format (not STR) to put into Excel
excel_date = date_stamp.strftime('%d.%m.%Y')
my_date_stamp = date_stamp.strftime('%y%m%d_%H-%M')
excel_sheet_date = date_stamp.strftime('%d %m %y')

thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
boldFont = Font(bold=True)


# Function that returns the letter of a column by column index: 1-->'A'; 2-->'B'; 3-->'C' etc.
def get_column_letter(column_index):
    col_letter = utils.cell.get_column_letter(column_index)
    return col_letter


# Function that returns all the important ranges for a table to write that to Excel:
# Range of the headers, e.g. "A1:Z1";
# Range of the body, e.g. "A1:Z7456";
# Range of the full chart, e.g. "A1:Z7456".
def chart_excel_ranges(chart):
    if type(chart) != list and type(chart[0]) != list:
        print("Check if your tables is a nested LIST -- should be! ")
    else:
        max_col = len(chart[0])
        max_col_letter = get_column_letter(max_col)
        max_row = len(chart)
        headers_range_str = f"A1:{max_col_letter}1"
        table_body_range_str = f"A2:{max_col_letter}{max_row}"
        full_table_range_str = f"A1:{max_col_letter}{max_row}"
        table_range_dict = {
            "headers": headers_range_str,
            "body": table_body_range_str,
            "full_table": full_table_range_str
        }
    return table_range_dict


# the function that returns a DICT with column letters ('A','B','C'...) & corresponding column width lengths as values
# Calculations are based on average lengths of data within a certain column. Also we use a DIVIDER in order to have a
# good looking table with same widths where possible.
# If the divider is 5, then column widths would be (10;15;5;15;20;25)
def width_len_dict(chart, min_width=1, max_width=50, width_divider=5):
    length_dict = {}
    col_count = len(chart[0])
    for column in range(col_count):
        temp_col_list = [len(str(my_row[column])) for my_row in chart if len(my_row)]
        values_col_list = [str(my_row[column]) for my_row in chart]
        header_width = temp_col_list[0]
        if header_width > min_width:
            min_width = header_width
        avg_length = int(sum(temp_col_list) / len(temp_col_list))
        if avg_length % width_divider != 0:
            avg_length = avg_length//width_divider
            avg_length = (avg_length+1) * width_divider
        if avg_length < min_width:
            avg_length = min_width
        elif avg_length > max_width:
            avg_length = max_width
        if "=HYPERLINK" in values_col_list[1]:
            avg_length = min_width
        if u"\u25AF" in values_col_list[1] or u"\u25AF" in values_col_list[1]:
            avg_length = max_width
        column_letter = utils.get_column_letter(column + 1)
        length_dict[column_letter] = avg_length+1
    return length_dict

# Function that formats a cell based on type of the value (int, float, str).
# Also this function checks the row of a cell:
# if the cell is in the 1st row than it is formatted as header: bold, color-filled, with borders

def random_color():
    random_color_key = random.choice(list(colors.keys()))
    random_color = colors[random_color_key]
    return random_color
headersColor = random_color()

def cell_format(input_cell, headers_color=headersColor):
    def cell_align(input_cell, align='center'):
        input_cell.alignment = Alignment(horizontal=align, vertical='center')
        return None
    cell_value = input_cell.value
    my_type = type(cell_value)
    if input_cell.row == 1:
        cell_align(input_cell, align='center')
        input_cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
        input_cell.fill = PatternFill('solid', fgColor=headers_color)
        input_cell.font = Font(bold=True)
        input_cell.font = Font(bold=True)
    elif type(cell_value) == int:
        input_cell.number_format = '#,##0'
    elif type(cell_value) == float and cell_value <= 1:
        input_cell.number_format = "0%"
    if my_type == int or my_type == float or (my_type == str and len(cell_value) < 11):
        cell_align(input_cell, align='left')
    if "=HYPERLINK" in str(cell_value):
        input_cell.style = 'Hyperlink'
        cell_align(input_cell, align='left')
    elif u"\u25AF" in str(cell_value) or u"\u25AE" in str(cell_value):
        input_cell.font = Font(size = 7)
        input_cell.alignment = Alignment(horizontal='left', vertical='center')
    return None

# Function that applies Excel auto-filter for the specific cell range, that is usually
# a header range: "A1:{max_col}1"
def excel_autofilter(worksheet_name, cell_range_str):
    worksheet_name.auto_filter.ref = cell_range_str
    return None


# Function that takes a dict as an input with sheetname as key, output chart as value and writes that down
# in a good format into a new excel document and saves it
def excel_multi_writer(output_charts_dict, loadWorkbook=None):
    if not loadWorkbook:
        wb = Workbook()
        existing_sheets = len(wb.sheetnames)
        needed_sheets = len(output_charts_dict)
        while existing_sheets < needed_sheets:
            wb.create_sheet()
            existing_sheets += 1
        sheet_index = 0
        for chart in output_charts_dict:
            ws = wb[wb.sheetnames[sheet_index]]
            ws.title = chart
            sheet_index+=1
    else:
        wb = loadWorkbook
        for sheet_name in output_charts_dict:
            wb.create_sheet(sheet_name)
    for chart in output_charts_dict:
        ws = wb[chart]
        output_chart = output_charts_dict[chart]
        for row in output_chart:
            ws.append(row)
        ranges_str = chart_excel_ranges(output_chart)
        width_dict = width_len_dict(output_chart)
        excel_autofilter(ws, ranges_str["headers"])
        for row in ws[ranges_str["full_table"]]:
            for cell in row:
                cell_format(cell)
        for col in width_dict:
            ws.column_dimensions[col].width = width_dict[col]
    return wb


def ws_conditional_formatting(conditions_dict, ws):
    for condition in conditions_dict:
        my_range = conditions_dict[condition][0]
        formula = condition
        color = conditions_dict[condition][1]
        color = PatternFill(bgColor=color)
        ws.conditional_formatting.add(
            my_range,
            FormulaRule(
                formula=[formula],
                stopIfTrue=True,
                fill=color,
                font=Font(bold=False)))
    return None

# conditions = {
#     'AND(B1=0,B1<>"")': [f"B1:B20000", zeroFill],
#     'B1=1': [f"B1:B20000", oneFill],
#     'B1=2': [f"B1:B20000", twoFill],
#     'B1=3': [f"B1:B20000", threeFill],
#     '$I2>5999': [f"C2:L20000", riskFill],
#     'AND($I2>3500,$I2<6000)': [f"C2:L20000", medRiskFill],
# }

# outputChartsDict = {
#     sheet_name: test_chart,
#     sheet_name2: test_chart2
# }

# wb = excel_multi_writer(outputChartsDict)
#
# for ws_name in wb.sheetnames:
#     ws = wb[ws_name]
#     ws_conditional_formatting(conditions, ws)
#
# wb.save(outputFullPath)

# ranges1 = chart_excel_ranges(test_chart)
# ranges2 = chart_excel_ranges(test_chart2)
